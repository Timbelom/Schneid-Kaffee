import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
import os
from tkinter import Tk
from tkinter.filedialog import askdirectory

# Function to prompt the user to select a folder
def select_folder():
    root = Tk()
    root.withdraw()  # Hide the root window
    folder_selected = askdirectory()  # Open folder selection dialog
    root.destroy()
    return folder_selected

# Function to sanitize sheet titles
def sanitize_sheet_title(title):
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        title = title.replace(char, '_')
    return title

# Function to read roast names from a text file
def read_roast_names(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        roast_names = [line.strip() for line in file if line.strip()]
    return roast_names

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Path to the roast_names.txt file
roast_names_file = os.path.join(script_dir, 'roast_names.txt')

# Read roast names from the text file
roast_names = read_roast_names(roast_names_file)

# Prompt user to select a folder
folder_path = select_folder()

# Extract the name of the selected folder
folder_name = os.path.basename(folder_path)

# Initialize the Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet created

# Define your file paths based on the selected folder and its name
file_paths = [os.path.join(folder_path, f'{month} {folder_name}.xlsx') for month in [
    'Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November', 'Dezember'
]]

id = 1

monthly_totals = {}
for file_path in file_paths:
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        df_reversed = df.iloc[::-1]
        new_rows = []
        weight_before_sum = 0
        weight_after_sum = 0
        last_date = None
        monthly_before_sum = 0
        monthly_after_sum = 0
        monthly_before_sum_check = 0
        monthly_after_sum_check = 0

        for index, row in df_reversed.iterrows():
            
            date = pd.to_datetime(row.iloc[1]).date()
            roast_title = row.iloc[3]
            weight_before = row.iloc[4]
            weight_after = row.iloc[5]
            monthly_before_sum += weight_before
            monthly_after_sum += weight_after
            name_found = ''
            text = ''

            for name in roast_names:
                if name in roast_title:
                    name_found = name
                    text = roast_title.split(name)[1].split(':')[0].strip() if ':' in roast_title else ''
                    break
            if not name_found:
                text = roast_title.split(':')[0]

            if date != last_date and last_date is not None:
                new_rows[-1].update({'Rohk.': weight_before_sum, 'Röstk.': weight_after_sum})
                weight_before_sum = 0
                weight_after_sum = 0
            
            if date != last_date:
                last_date = date
            else:
                date = ''
            
            weight_before_sum += weight_before
            weight_after_sum += weight_after
            if date != '':
                date_obj = datetime.strptime(str(date), "%Y-%m-%d")
                date = date_obj.strftime("%d.%m.%Y")
            new_rows.append({
                'ID': id,
                'Datum': str(date),
                'Name': name_found,
                'Sorte': text,
                'Rohk..': weight_before,
                'Rohk.': '',
                'Röstk..': weight_after,
                'Röstk.': ''
            })
            id += 1
            monthly_before_sum_check += weight_before
            monthly_after_sum_check += weight_after
        if new_rows:
            new_rows[-1].update({'Rohk.': weight_before_sum, 'Röstk.': weight_after_sum})

        new_df = pd.DataFrame(new_rows)
        title = sanitize_sheet_title(os.path.splitext(os.path.basename(file_path))[0])
        ws = wb.create_sheet(title=title)  # Create a sheet named after the month
        ws.freeze_panes = 'A3' 
        # Insert the total row at the top
        ws.merge_cells('A1:C1')
        cell = ws['A1']
        cell.value = f'Röstbuch Schneid {title}'
        cell.alignment = Alignment(horizontal='center')
        font = Font(size=12, bold=True)
        cell.font = font
        ws['D1'] = 'Gesamt:'
        ws['E1'] = monthly_before_sum
        ws['G1'] = monthly_after_sum
        ws['F1'] = monthly_before_sum_check
        ws['H1'] = monthly_after_sum_check

        monthly_totals[title] = (monthly_before_sum, monthly_after_sum)

        thin_border = Border(bottom=Side(style='thin'))

        for col, header in enumerate(new_df.columns, start=1):
            ws.cell(row=2, column=col, value=header)

        max_length = {}
        for rowIndex, row in enumerate(new_df.itertuples(), start=3):  # Adjust start for new row index
            for colIndex, cell_value in enumerate(row[1:], start=1):  # row is a tuple including index at first position
                cell = ws.cell(row=rowIndex, column=colIndex, value=cell_value)
                if colIndex == 2 and getattr(row, 'Datum') != '':
                    for i in range(1, len(new_df.columns) + 1):
                        ws.cell(row=rowIndex - 1, column=i).border = thin_border
                if cell_value and isinstance(cell_value, str):
                    max_length[colIndex] = max(max_length.get(colIndex, 0), len(cell_value))
                if colIndex in [5, 6, 7, 8] and cell_value != '':
                    cell.number_format = '0.0'

        for i, width in max_length.items():
            ws.column_dimensions[get_column_letter(i)].width = width + 2
        print(f'Blatt für {title} populiert\n')
    else:
    # File does not exist
        title = sanitize_sheet_title(os.path.splitext(os.path.basename(file_path))[0])
        print(f"Datei für {title} kann nicht gefunden werden")

summary_ws = wb.create_sheet(title=sanitize_sheet_title(f"Röstk.Ges.{folder_name}"), index=0)
summary_ws.append(["Monat", "Rohk", "Röstk"])

max_length = [0, 0, 0]
for month, totals in monthly_totals.items():
    summary_ws.append([month, totals[0], totals[1]])
    max_length[0] = max(max_length[0], len(month))
    max_length[1] = max(max_length[1], len(str(totals[0])))
    max_length[2] = max(max_length[2], len(str(totals[1])))

for i, width in enumerate(max_length, start=1):
    summary_ws.column_dimensions[get_column_letter(i)].width = width + 2

for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=2, max_col=3):
    for cell in row:
        cell.number_format = '0.0'

print('Zusammenfassungsblatt erstellt\n')
wb.save(f'Röstbuch {folder_name}.xlsx')
print('Erfolg!')
