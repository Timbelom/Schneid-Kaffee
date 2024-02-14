import logging
import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
import openpyxl
import os
import calendar
import tkinter as tk
from tkinter import ttk

logging.basicConfig(filename='errorlog.log', level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')


class Employee:
    def __init__(self, employee_id, name):
        self.employee_id = employee_id
        self.name = name
        self.shifts ={}
        
    def add_shift(self, date, in_time, out_time, total_time):
        self.shifts[date] = [in_time, out_time, total_time]
        
def list_employees(api_key):
    # API endpoint for listing employees
    url = "https://api.mytimestation.com/v1.2/employees"

    # Make the GET request with HTTP Basic Authentication
    response = requests.get(url, auth=HTTPBasicAuth(api_key, ''))

    # Create and return a list of Employee objects if the request is successful
    if response.status_code == 200:
        employees_data = response.json()
        return [Employee(employee['employee_id'], employee['name']) for employee in employees_data['employees']]
    else:
        logging.error("Failed to retrieve employees. Status code: %s", response.status_code)
        print("Failed to retrieve employees. Status code:", response.status_code)
        return []

def get_hours_worked_report(api_key, employees, start_date, end_date):
    # API endpoint for fetching shifts
    base_url = "https://api.mytimestation.com/v1.2/shifts"
    
    def minutes_to_hhmm(total_minutes):
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    
    def convert_datetime_format(date_string):
    # Parse the string into a datetime object
        dt = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S")
        
        # Format the datetime object into the desired format
        formatted_date = dt.strftime("%m/%d/%Y %H:%M")
        
        return formatted_date

    # Loop through each Employee object and fetch their shifts
    for employee in employees:
        # Construct the full URL with parameters
        full_url = f"{base_url}?employee_id={employee.employee_id}&start_date={start_date}&end_date={end_date}"

        # Make the GET request with HTTP Basic Authentication
        response = requests.get(full_url, auth=HTTPBasicAuth(api_key, ''))

        # Check if the request was successful
        if response.status_code == 200:
            # Print the employee name and their shift data
            print(f"\nName: {employee.name}")
             # Extract and print total_minutes and in/out status for each shift
            shifts = response.json().get('shifts', [])
            for shift in shifts:
                # Formating the dict data
                
                total_minutes = shift.get('total_minutes')
                if total_minutes:
                    formatted_time = minutes_to_hhmm(int(total_minutes))
                
                in_status = shift.get('in')
                in_time = next(iter(in_status.items()))[1] if in_status else 'N/A'
                if in_time != 'N/A':
                    in_time = in_time[:-6]
                    
                out_status = shift.get('out')
                if out_status:
                    out_time = next(iter(out_status.items()))[1] if in_status else 'N/A'
                    if out_time != 'N/A':
                        out_time = out_time[:-6]
                    
                # there's a better way to do this but this works
                date =convert_datetime_format(in_time[:-9]+" "+in_time[11:])[:-6]
                clockin = convert_datetime_format(in_time[:-9]+" "+in_time[11:])[11:]
                if out_status:
                    clockout = convert_datetime_format(out_time[:-9]+" "+out_time[11:])[11:] 
                    employee.add_shift(date,clockin,clockout,formatted_time)
                else:
                    employee.add_shift(date,clockin,'N/A','N/A')
                
            for shift in employee.shifts:
                print(shift, employee.shifts[shift])
        else:
            print(f"Failed to retrieve shifts for employee {employee.name}. Status code:", response.status_code)


def fill_missing_dates(employee,start_date):
    
    def convert_datetime_format(date_string):
    # Parse the string into a datetime object
        dt = datetime.strptime(date_string, "%Y-%m-%d")
        # Format the datetime object into the desired format
        formatted_date = dt.strftime("%m/%d/%Y")
        return formatted_date
    
    # Convert start and end dates to datetime objects
    start = datetime.strptime(convert_datetime_format(start_date), "%m/%d/%Y")
    end = datetime.strptime(max(employee.shifts.keys()), "%m/%d/%Y")

    # Initialize a date variable with the start date
    current_date = start

    # Iterate over the range of dates
    while current_date <= end:
        # Format the current date as MM/DD/YYYY
        date_str = current_date.strftime("%m/%d/%Y")

        # If the date is not in the dictionary, add it with a default value
        if date_str not in employee.shifts:
            employee.shifts[date_str] = [0, 0, 0]

        # Move to the next day
        current_date += timedelta(days=1)
    def sort_shift_data(shift_data):
        # Sort the dictionary by date keys in ascending order
        sorted_data = dict(sorted(shift_data.items(), key=lambda item: datetime.strptime(item[0], "%m/%d/%Y")))
        return sorted_data
    employee.shifts=sort_shift_data(employee.shifts)


def find_and_open_excel_files(employee ,folder_path):
    
    def get_month_name(date_str):
        # Parse the date string into a datetime object
        date_obj = datetime.strptime(date_str, '%m/%d/%Y')

        # Extract the month name and year
        month_name = date_obj.strftime('%b')
        year = date_obj.strftime('%y')

        # Return the formatted string
        return f"{month_name}.{year}"
    # List all .xlsx files in the given folder
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    search_string = employee.name
    
    sheet_name = get_month_name(next(iter(employee.shifts)))

    # Check each file to see if it contains the search string in its name
    fileFound=False#Dumb perfectionism
    for file in excel_files:
        if search_string in file:
            # Construct the full file path
            file_path = os.path.join(folder_path, file)

            # Open the workbook - add your logic here as needed
            workbook = openpyxl.load_workbook(file_path)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
            sheet = workbook[sheet_name]

            # Start from row 7, column 3 (which is 'C7')
            row = 7
            date_column = 1
            in_column = 2
            out_column = 3

            # Iterate over the data list and write to the cells
            for shift in employee.shifts:
                date_cell = sheet.cell(row=row, column=date_column)
                date_cell.value = shift
                in_cell = sheet.cell(row=row, column=in_column)
                in_cell.value = employee.shifts[shift][0]
                out_cell = sheet.cell(row=row, column=out_column)
                out_cell.value = employee.shifts[shift][1]
                row += 1  # Move to the next row

            # Save the workbook
            workbook.save(file_path)

            # Example: Print the sheet names
            print(f"Updated file for {employee.name}")
            fileFound=True#Dumb perfectionism
            break
    if fileFound != True:
        print(f"No file found for {employee.name}")


def get_first_and_last_day_of_month():
    # Display year selection
    year = input("Enter a year (e.g., 2022): ")
    
    # Display month selection
    print("Select a month:")
    months = ["1. January", "2. February", "3. March", "4. April", "5. May", "6. June", 
              "7. July", "8. August", "9. September", "10. October", "11. November", "12. December"]
    for month in months:
        print(month)
    month_choice = input("Enter the number of the month: ")
    
    # Validate the user's input and return the first and last day of the month
    try:
        year = int(year)
        month_choice = int(month_choice)
        if month_choice < 1 or month_choice > 12:
            raise ValueError("Invalid month number.")
        
        # Create a datetime object for the first day of the selected month
        first_day_date = datetime(year, month_choice, 1)
        first_day_str = first_day_date.strftime("%Y-%m-%d")
        
        # Find the last day of the month
        last_day = calendar.monthrange(year, month_choice)[1]
        last_day_date = datetime(year, month_choice, last_day)
        last_day_str = last_day_date.strftime("%Y-%m-%d")
        
        # Return the first and last day of the month
        return first_day_str, last_day_str
    except ValueError as e:
        print(f"Invalid input: {e}. Please enter numeric values only.")
        return None, None

    
def submit_action():
    year = year_var.get()
    month = month_var.get()
    
    # Calculate the first and last day of the selected month
    first_day_date = datetime(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    last_day_date = datetime(year, month, last_day)
    
    # Store the dates in the global variable
    date_result["first_day"] = first_day_date.strftime('%Y-%m-%d')
    date_result["last_day"] = last_day_date.strftime('%Y-%m-%d')
    
    # Close the window
    root.destroy()
# Main execution
if __name__ == "__main__":
    try:
        # API key should be saved in a separate .env file stored locally
        with open("APIKEY.env", 'r') as file:
            api_key = file.read()

        # Calculate the first day of the current month
            
        date_result = {"first_day": "", "last_day": ""}

        # Create the main window
        root = tk.Tk()
        root.title("Date Selection")

        # Variables for dropdown selections
        year_var = tk.IntVar(value=datetime.now().year)  # Default to current year
        month_var = tk.IntVar(value=datetime.now().month)  # Default to current month

        # Create Year dropdown
        year_label = ttk.Label(root, text="Select Year:")
        year_label.pack()
        year_dropdown = ttk.Combobox(root, textvariable=year_var, width=15)
        year_dropdown['values'] = tuple(range(2023, 2101))  # Example range from 1900 to 2100
        year_dropdown.pack()

        # Create Month dropdown
        month_label = ttk.Label(root, text="Select Month:")
        month_label.pack()
        month_dropdown = ttk.Combobox(root, textvariable=month_var, width=15)
        month_dropdown['values'] = tuple(range(1, 13))  # Months 1 to 12
        month_dropdown.pack()

        # Submit button
        submit_btn = ttk.Button(root, text="Submit", command=submit_action)
        submit_btn.pack()

        # Run the application
        root.mainloop()
            
        # start_date, end_date = get_first_and_last_day_of_month()

        start_date = date_result["first_day"]
        end_date = date_result["last_day"]
        # today = datetime.now()
        # start_date = today.replace(day=1).strftime("%Y-%m-%d")
        # start_date = "2024-01-01"

        # End date is today's date
        # end_date = today.strftime("%Y-%m-%d")
        # end_date= "2024-01-31"

        # Get the list of Employee objects
        employees = list_employees(api_key)

    # Get the hours worked report for each employee
    if employees:
        get_hours_worked_report(api_key, employees, start_date, end_date)
        for employee in employees:
            if employee.shifts:
                fill_missing_dates(employee, start_date)
                find_and_open_excel_files(employee, "timestation")
        print("Success!")
